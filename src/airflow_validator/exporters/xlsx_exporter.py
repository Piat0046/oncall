"""XLSX 파일 생성"""

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet


class XLSXExporter:
    """XLSX 내보내기"""

    def __init__(self, output_dir: str = "./output"):
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # 스타일 정의
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.failed_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )
        self.warning_fill = PatternFill(
            start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
        )
        self.center_align = Alignment(horizontal="center", vertical="center")

    def export(
        self,
        failure_data: list[dict[str, Any]],
        sync_failures: list[dict[str, Any]],
        target_date: str,
    ) -> str:
        """
        실패 정보를 XLSX 파일로 내보내기

        Args:
            failure_data: Airflow 실패 정보 목록
            sync_failures: MySQL 동기화 실패 데이터
            target_date: 대상 날짜 문자열

        Returns:
            생성된 파일 경로
        """
        wb = Workbook()

        # Sheet 1: 동기화 실패 데이터 (MySQL)
        ws_sync = wb.active
        ws_sync.title = "Sync Failures"
        self._create_sync_failures_sheet(ws_sync, sync_failures)

        # Sheet 2: DAG Run Summary (Airflow)
        ws_summary = wb.create_sheet("DAG Run Summary")
        self._create_summary_sheet(ws_summary, failure_data)

        # Sheet 3: Failed Tasks (Airflow)
        ws_tasks = wb.create_sheet("Failed Tasks")
        self._create_tasks_sheet(ws_tasks, failure_data)

        # Sheet 4: DAG Configs (Airflow)
        ws_config = wb.create_sheet("DAG Configs")
        self._create_config_sheet(ws_config, failure_data)

        # 파일 저장
        timestamp = datetime.now().strftime("%H%M%S")
        filename = f"airflow_failures_{target_date}_{timestamp}.xlsx"
        filepath = self.output_dir / filename
        wb.save(filepath)

        return str(filepath)

    def _create_sync_failures_sheet(
        self, ws: Worksheet, sync_failures: list[dict[str, Any]]
    ):
        """동기화 실패 데이터 시트 생성 (MySQL 쿼리 결과)"""
        headers = [
            "Provider",
            "User ID",
            "Data ID",
            "Data Category",
            "Mall ID",
            "Mall Name",
            "Data Set Name",
            "Updated At",
            "Plan End Date",
        ]

        # 헤더 작성
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border

        # 데이터 작성
        for row_idx, row in enumerate(sync_failures, 2):
            ws.cell(row=row_idx, column=1, value=row.get("provider", ""))
            ws.cell(row=row_idx, column=2, value=row.get("user_id", ""))
            ws.cell(row=row_idx, column=3, value=row.get("data_id", ""))
            ws.cell(row=row_idx, column=4, value=row.get("data_category", ""))
            ws.cell(row=row_idx, column=5, value=row.get("mall_id", ""))
            ws.cell(row=row_idx, column=6, value=row.get("mall_name", ""))
            ws.cell(row=row_idx, column=7, value=row.get("data_set_name", ""))
            ws.cell(
                row=row_idx,
                column=8,
                value=self._format_datetime(row.get("updated_at")),
            )
            ws.cell(
                row=row_idx,
                column=9,
                value=self._format_datetime(row.get("plan_end_consumer")),
            )

            # 스타일 적용
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = self.border
                cell.fill = self.warning_fill

        self._auto_adjust_column_width(ws)

    def _create_summary_sheet(
        self, ws: Worksheet, failure_data: list[dict[str, Any]]
    ):
        """DAG Run 요약 시트 생성"""
        headers = [
            "DAG ID",
            "Run ID",
            "Execution Date",
            "State",
            "Run Type",
            "Start Date",
            "End Date",
            "Failed Task Count",
        ]

        # 헤더 작성
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border

        # 데이터 작성
        for row_idx, run in enumerate(failure_data, 2):
            ws.cell(row=row_idx, column=1, value=run.get("dag_id", ""))
            ws.cell(row=row_idx, column=2, value=run.get("run_id", ""))
            ws.cell(
                row=row_idx,
                column=3,
                value=self._format_datetime(run.get("execution_date")),
            )
            ws.cell(row=row_idx, column=4, value=run.get("state", ""))
            ws.cell(row=row_idx, column=5, value=run.get("run_type", ""))
            ws.cell(
                row=row_idx,
                column=6,
                value=self._format_datetime(run.get("start_date")),
            )
            ws.cell(
                row=row_idx,
                column=7,
                value=self._format_datetime(run.get("end_date")),
            )
            ws.cell(row=row_idx, column=8, value=run.get("failed_task_count", 0))

            # 스타일 적용
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = self.border
                cell.fill = self.failed_fill

        self._auto_adjust_column_width(ws)

    def _create_tasks_sheet(
        self, ws: Worksheet, failure_data: list[dict[str, Any]]
    ):
        """실패한 Task 상세 시트 생성"""
        headers = [
            "DAG ID",
            "Run ID",
            "Task ID",
            "Operator",
            "Duration (sec)",
            "Try Number",
            "Hostname",
        ]

        # 헤더 작성
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border

        # 데이터 작성
        row_idx = 2
        for run in failure_data:
            dag_id = run.get("dag_id", "")
            run_id = run.get("run_id", "")

            for task in run.get("failed_tasks", []):
                ws.cell(row=row_idx, column=1, value=dag_id)
                ws.cell(row=row_idx, column=2, value=run_id)
                ws.cell(row=row_idx, column=3, value=task.get("task_id", ""))
                ws.cell(row=row_idx, column=4, value=task.get("operator", ""))
                ws.cell(row=row_idx, column=5, value=task.get("duration"))
                ws.cell(row=row_idx, column=6, value=task.get("try_number"))
                ws.cell(row=row_idx, column=7, value=task.get("hostname", ""))

                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).border = self.border

                row_idx += 1

        self._auto_adjust_column_width(ws)

    def _create_config_sheet(
        self, ws: Worksheet, failure_data: list[dict[str, Any]]
    ):
        """Config 정보 시트 생성"""
        headers = ["DAG ID", "Run ID", "Config Key", "Config Value"]

        # 헤더 작성
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border

        # 데이터 작성
        row_idx = 2
        for run in failure_data:
            dag_id = run.get("dag_id", "")
            run_id = run.get("run_id", "")
            conf = run.get("conf", {})

            if conf:
                for key, value in conf.items():
                    ws.cell(row=row_idx, column=1, value=dag_id)
                    ws.cell(row=row_idx, column=2, value=run_id)
                    ws.cell(row=row_idx, column=3, value=str(key))
                    ws.cell(row=row_idx, column=4, value=self._format_value(value))

                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row_idx, column=col).border = self.border

                    row_idx += 1
            else:
                # Config가 없는 경우
                ws.cell(row=row_idx, column=1, value=dag_id)
                ws.cell(row=row_idx, column=2, value=run_id)
                ws.cell(row=row_idx, column=3, value="(no config)")
                ws.cell(row=row_idx, column=4, value="-")

                for col in range(1, len(headers) + 1):
                    ws.cell(row=row_idx, column=col).border = self.border

                row_idx += 1

        self._auto_adjust_column_width(ws)

    def _auto_adjust_column_width(self, ws: Worksheet):
        """열 너비 자동 조정"""
        for column_cells in ws.columns:
            max_length = 0
            column_letter = column_cells[0].column_letter

            for cell in column_cells:
                try:
                    cell_length = len(str(cell.value or ""))
                    if cell_length > max_length:
                        max_length = cell_length
                except Exception:
                    pass

            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    @staticmethod
    def _format_datetime(value: datetime | str | None) -> str:
        """datetime을 문자열로 포맷"""
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S")
        return str(value)

    @staticmethod
    def _format_value(value: Any) -> str:
        """값을 문자열로 포맷"""
        if isinstance(value, (dict, list)):
            return json.dumps(value, ensure_ascii=False, default=str)
        return str(value)
